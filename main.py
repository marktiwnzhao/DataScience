from gensim.models import Word2Vec, KeyedVectors
import pandas as pd
import gensim
import numpy as np
from scipy import spatial
import time
from RenewDict import RenewDict
import Accuracy
from Calculate import avg_feature_vector
import pickle
from openpyxl import load_workbook
from SM import findTopk_SM, findTop1_SM


def open_excel(path, model, set_, line_name):
    f = pd.read_excel(path, sheet_name=1, usecols=line_name)
    data = {}
    for line in range(len(f)):
        this_str = format(f.loc[line].values[0])
        vector = avg_feature_vector(sentence=this_str, model=model, num_features=200, index2word_set=set_)
        if np.any(vector):
            data[this_str] = vector
    return data


# 加载txt词向量文件, 并将其保存为二进制文件 https://blog.csdn.net/orangerfun/article/details/120253144
def load_word_vec(vec_path):
    wv_from_text = gensim.models.KeyedVectors.load_word2vec_format(vec_path, binary=False)
    # 如果每次都用上面的方法加载，速度非常慢，可以将词向量文件保存成bin文件，以后就加载bin文件，速度会变快
    wv_from_text.init_sims(replace=True)
    wv_from_text.save(vec_path.replace(".txt", ".bin"))


if __name__ == '__main__':

    # load_word_vec("tencent-ailab-embedding-zh-d200-v0.2.0-s.txt")
    embed_path = "tencent-ailab-embedding-zh-d200-v0.2.0-s.bin"
    filename = "字段关联关系.xlsx"

    # 创建RenewDict类，实现动态加载数据元并更新字典
    renew = RenewDict(embed_path, filename)

    wv_from_text = KeyedVectors.load(embed_path, mmap='r')
    index2word_set = set(wv_from_text.index_to_key)
    with open('source.pkl', 'rb') as f:
        source = pickle.load(f)

    # 计算Accuracy
    # wv_from_text = KeyedVectors.load(embed_path, mmap='r')
    # index2word_set = set(wv_from_text.index_to_key)
    # Accuracy.accuracy(wv_from_text, index2word_set)

    # 输入query和k，输入“quit”结束循环
    while True:
        query = input()
        if query == "quit":
            break
        if query == "add":
            # 输入要增加的数据，四列的数据元之间以空格隔开
            data = input().split()
            target = [(data[0], data[1], data[2], data[3])]
            df = pd.DataFrame(target, columns=['机构名称', '表中文名', '字段中文（可忽略）', '数据元'])  # 列表数据转为数据框
            df1 = pd.DataFrame(pd.read_excel(filename, sheet_name=1))  # 读取原数据文件和表
            book = load_workbook(filename)
            writer = pd.ExcelWriter(filename, engine='openpyxl')  # 必须先load_workbook再writer
            writer.book = book
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
            df_rows = df1.shape[0]  # 获取原数据的行数
            df.to_excel(writer, sheet_name="关联关系", startrow=df_rows + 1, index=False,
                        header=False)  # 将数据写入excel中的关联关系表,从第一个空行开始写
            writer.save()  # 保存
            renew.renew()
            continue
        k = int(input())
        test_vector = avg_feature_vector(query, wv_from_text, 200, index2word_set)
        my_dict = {}
        for i, (key, value) in enumerate(source.items()):
            sim = 1 - spatial.distance.cosine(test_vector, value)
            my_dict[key] = sim
        test_data_2 = sorted(my_dict.items(), key=lambda x: x[1], reverse=True)
        print("Using Method 1 : Average feature vector")
        print(test_data_2[0:k])
        print("Using Method 2 : SequenceMatcher")
        # findTopk_SM(query, k)
        findTopk_SM(query, k)

    # 直接比较
    # test_vector = avg_feature_vector("姓名", wv_from_text, 200, index2word_set)
    # k = 5
    # print(time.time() - t0)
